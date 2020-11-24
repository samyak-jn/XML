from app import app
from flask import request, redirect
from flask.templating import render_template
import urllib
from werkzeug.utils import redirect
from werkzeug.utils import secure_filename
import os
import xml.etree.ElementTree as ET
import pandas as pd
from flask.helpers import flash, send_file, send_from_directory, url_for
from lxml import etree
import calendar
import time
from openpyxl import Workbook
import re
import openpyxl
from lxml import etree as et
import math
import io

app.secret_key = "super secret key"
app.config["ALLOWED_FILE_EXTENSIONS"] = {"XML", "CSV", "XLSX", "xlsx"}
app.config['MAX_CONTENT_LENGTH'] = 700 * 1024 * 1024
xmlDocument = r'instance/uploads/'

def get_namespace(element):
    m = re.match('\{.*\}', element.tag)
    return m.group(0) if m else ''

def allowed_file(filename):
    if not "." in filename:
        return False
    # for the extension
    ext = filename.rsplit(".", 1)[1]
    if ext.upper() in app.config["ALLOWED_FILE_EXTENSIONS"]:
        return True
    else:
        return False

def allowed_file_filesize(filesize):
    if int(filesize) <= app.config["MAX_FILE_FILESIZE"]:
        return True
    else:
        return False

def upload():
    if request.method == "POST":
        if request.files:
            if "filesize" in request.cookies:
                if not allowed_file_filesize(request.cookies["filesize"]):
                    print("Filesize exceeded maximum limit")
                    return redirect(request.url)
            file = request.files["file"]

            if file.filename == "":
                print("No filename")
                return redirect(request.url)
            if allowed_file(file.filename):
                filename = secure_filename(file.filename)
                os.makedirs(os.path.join(app.instance_path, 'uploads'), exist_ok=True)
                file.save(os.path.join(app.instance_path, 'uploads', secure_filename(filename)))
                print("File Saved")
                return redirect(url_for('upload_file',filename=filename))
            else:
                print("That file extension is not allowed.")
                return redirect(request.url)

xmlDocument = r'instance/uploads/'

def extract_local_tag(tagname):
    return tagname.split('}')[-1].strip()

def xml_to_dataframe(xmlDocument):
    class_data = []
    data = []

    for event,elem in ET.iterparse(xmlDocument, events=('start', 'end')):
        tag = extract_local_tag(elem.tag)
        if event=='start' and tag=='managedObject':
            class_data=[elem.get('class').strip(),elem.get('version').strip(),elem.get('distName').strip(),elem.get('id').strip()]
        if event=='start' and tag=='p':
            data.append(class_data+[elem.get('name'),elem.text])
    df = pd.DataFrame(data,columns=['class','version','distName','id','parameter','value'])

    return df

def updateXML(xmlDocument,class_=None,sites=None,param_dict=None,cells=None):
    param_for_list = {}
    if param_dict!=None:
        for k,v in param_dict.items():
            if '-' in k:
                param_for_list[k] = v

    tree = etree.parse(xmlDocument)
    root = tree.getroot().findall('*')[0]

    relevent = []
  
    for elem in tree.findall('//{raml20.xsd}managedObject'):
            site = elem.get('distName').split('/')[1].split('-')[1].strip()
            if class_ is not None:
                cell = elem.get('distName').lower().split(class_+'-')[-1]
            else:
                cell = elem.get('distName').lower().split('-')[-1]
       #     cell = elem.get('distName').lower().split(class_+'-')[-1]

            if ((class_ is None )or(elem.attrib['class'].strip().lower()== class_ )) and (((sites is None)or(site in 
sites))and((cells is None)or(cell in cells))):
                relevent.append(elem)
            else:
                root.remove(elem)
    if param_dict is not None:
        for elem in relevent:
                for p in elem.findall('{raml20.xsd}p'):
                    if(p.get('name').strip().lower() in param_dict):
                        p.text = param_dict.get(p.get('name').strip().lower())
                    else:
                        elem.remove(p)
                # For handling list items
                if len(param_for_list)==0:
                    for l in elem.findall('{raml20.xsd}list'):
                        elem.remove(l)
                for param,value in param_for_list.items():
                    items = param.split('-')
                    list_name = items[0].strip().lower()
                    item_name = items[2].strip().lower()
                    try:
                        item_number = int(items[1])
                    except(ValueError):
                        # case of all
                        item_number = items[1].strip().lower()


                    for i in elem.findall('{raml20.xsd}list'):

                        if i.get('name').strip().lower()==list_name:

                            # if a param from all items of a list need to be updated
                            if item_number == "all":
                                for item in i.findall('{raml20.xsd}item'):
                                    for p in item.findall('{raml20.xsd}p'):
                                        if (p.get('name').strip().lower() == item_name.strip().lower()):
                                            p.text = value
                                        if p.get('name').strip().lower() not in [x.split('-')[2].strip().lower() for x in list(param_for_list.keys())]:
                                            item.remove(p)

                            # If a particular index of item needs to be updated
                            else:
                                try:
                                    for p in (i.getchildren()[item_number-1].findall('{raml20.xsd}p')):
                                        if (p.get('name').strip().lower() == item_name.strip().lower()):
                                            p.text = value
                                        if p.get('name').strip().lower() not in [x.split('-')[2].strip().lower() for x in list(param_for_list.keys())]:
                                            i.getchildren()[item_number-1].remove(p)
                                except(IndexError):
                                    # Remove list if item number is wrong 
                                    print('Index Error for list name:{}'.format(i.get('name')))
                        if (i.get('name').strip().lower() not in [x.split('-')[0].strip().lower() for x in list(param_for_list.keys())]):
                            elem.remove(i)
    et = etree.ElementTree(tree.getroot())
   # print(etree.tostring(tree,encoding="unicode", pretty_print=True))
    et.write('app/download/download.xml', pretty_print=True)
    return

def bulkupdateXML(xmlDocument, inputDocument):
    df = pd.read_csv(inputDocument)
    tree = etree.parse(xmlDocument)
    root =  tree.getroot().findall('*')[0]
    data = df.values
    data[:,1] = [i.strip().lower() for i in data[:,1]]
    data[:,2] = [i.strip().lower() for i in data[:,2]]
    data[:,0] = [i.strip().split(',') for i in data[:,0]]
    for elem in tree.findall('//{raml20.xsd}managedObject'):
            site = elem.get('distName').split('/')[1].split('-')[1].strip()
            flag = False
            for row in data:
                class_ = row[1]
                sites = row[0]
                param_list = {row[2]:str(row[3])}
                if elem.attrib['class'].strip().lower()== class_ and (site in sites) :
                    flag = True   
                    for i in elem.findall('*'):
                        if i.tag!='{raml20.xsd}p':
                            elem.remove(i)
                    for p in elem.findall('{raml20.xsd}p'):
                        if (p.get('name').strip().lower() in param_list):
                            p.text = param_list.get(p.get('name').strip().lower())
                        elif (p.get('name').strip().lower() not in data[data[:,1]==class_][:,2]):
                            elem.remove(p)
            if flag==False:
                root.remove(elem)
    et = etree.ElementTree(tree.getroot())
    #print(etree.tostring(tree,encoding="unicode", pretty_print=True))
    et.write('app/download/download.xml', pretty_print=True)
    return


def dumpparser(filepath):
    parameter_tracker = {}
    rowcol_tracker = {}
    visited_class = []
    currently_active_sheet = ''
    dt_string = calendar.timegm(time.gmtime())
    dest_filename = str('dump_' + str(dt_string) + '.xlsx')
    wb = Workbook()
    ws = ''
    context = etree.iterparse(filepath, events=('start', 'end'))
    for event,root in context:
        namespace = get_namespace(root)
        if event == "end" and root.tag == str(namespace + 'managedObject'):
            classname = root.attrib['class']
            if classname not in visited_class:
                visited_class.append(classname)
                ws = wb.create_sheet(title=classname)
                currently_active_sheet = classname
                parameter_tracker[classname] = {}
                rowcol_tracker[classname] = {}
                rowcol_tracker[classname]['row'] = 1
                rowcol_tracker[classname]['col'] = 0
            if currently_active_sheet != classname:
                ws = wb[classname]
                currently_active_sheet = classname
            rowcol_tracker[classname]['row'] = rowcol_tracker[classname]['row'] + 1
            distName = root.attrib['distName']
            for d in distName.split('/'):
                dn = d.split('-')
                if dn[0] != 'PLMN':
                    if dn[0] not in parameter_tracker[classname]:
                        rowcol_tracker[classname]['col'] = rowcol_tracker[classname]['col'] + 1
                        parameter_tracker[classname][dn[0]] = {}
                        parameter_tracker[classname][dn[0]] = rowcol_tracker[classname]['col']
                        ws.cell(1, rowcol_tracker[classname]['col'], dn[0])
                        ws.cell(rowcol_tracker[classname]['row'], rowcol_tracker[classname]['col'], dn[1])
                    else:
                        ws.cell(rowcol_tracker[classname]['row'], parameter_tracker[classname][dn[0]], dn[1])
            for p in root.findall(str(namespace + 'p')):
                pname = p.attrib['name']
                if pname not in parameter_tracker[classname]:
                    rowcol_tracker[classname]['col'] += 1
                    parameter_tracker[classname][pname] = {}
                    parameter_tracker[classname][pname] = rowcol_tracker[classname]['col']
                    ws.cell(1, rowcol_tracker[classname]['col'],pname)
                ws.cell(rowcol_tracker[classname]['row'],parameter_tracker[classname][pname],p.text)
            for list in root.findall(str(namespace + 'list')):
                listname = list.attrib['name']
                itemcount = 0
                pcount = 0
                for item in list.findall(str(namespace + 'item')):
                    itemcount = itemcount + 1
                    for p in item.findall(str(namespace + 'p')):
                        pname = p.attrib['name']
                        parra = str(listname + ':' + 'item' + str(itemcount) + ':' + pname)
                        if parra not in parameter_tracker[classname]:
                            rowcol_tracker[classname]['col'] += 1
                            parameter_tracker[classname][parra] = {}
                            parameter_tracker[classname][parra] = rowcol_tracker[classname]['col']
                            ws.cell(1, rowcol_tracker[classname]['col'], parra)
                        ws.cell(rowcol_tracker[classname]['row'], parameter_tracker[classname][parra], p.text)
                for p in list.findall(str(namespace + 'p')):
                    pcount +=1
                    pconcat = str(listname + ':p' + str(pcount))
                    if pconcat not in parameter_tracker[classname]:
                        rowcol_tracker[classname]['col'] += 1
                        parameter_tracker[classname][pconcat] = {}
                        parameter_tracker[classname][pconcat] = rowcol_tracker[classname]['col']
                        ws.cell(1, rowcol_tracker[classname]['col'], pconcat)
                    ws.cell(rowcol_tracker[classname]['row'], parameter_tracker[classname][pconcat], p.text)
        root.clear
    wb.save(filename='instance/uploads/dump.xlsx')

def filter_dump(filter_input,dump):
    
    # making dataframe of all sheets
    df_class = pd.read_excel(filter_input, sheet_name = 'Class')
    df_siteID = pd.read_excel(filter_input, sheet_name='SiteID')
    df_param = pd.read_excel(filter_input, sheet_name='Parameters')
    #extracting class, site_id and parameters to be filtered
    class_list = df_class['Class'].tolist()
    siteid_list1 = df_siteID['SiteID'].tolist()
    siteid_list = list(map(str, siteid_list1)) # converting site id to string
    param_list = df_param['parameter'].tolist()
    param_list = [i.lower() for i in param_list]
    wb = openpyxl.load_workbook(dump)
    sheet_names = wb.sheetnames
    #deleting irrelevent sheets in dump
    for i in sheet_names:
        if i not in class_list:
            del wb[i]
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        r = 2
        while r <= ws.max_row:
            if ws.cell(row=r, column=1).value not in siteid_list:
                ws.delete_rows(r, 1)
            else:
                r += 1
        c = 1
        while c <= ws.max_column:
            if (ws.cell(row=1, column=c).value.lower() not in param_list) and not ws.cell(row=1, column=c).value.isupper():
                ws.delete_cols(c)
            else:
                c += 1
        c = 1
        while c <= ws.max_column:
            if ws.cell(row=1, column=c).value.isupper():
                c += 1
            else:
                ws.insert_cols(c)
                ws.cell(row=1,column=c).value = 'operation'
                break
    wb.save(filename="app/download/download.xlsx")

def create_XML(updated_filter,xmlDocument):
    
    wb = openpyxl.load_workbook(updated_filter)
    sheet_names = wb.sheetnames
    
    
    # getting the XML tree structure
    tree = et.parse(xmlDocument)
    
    #extracting root of xml
    root =  tree.getroot().findall('*')[0]

    namespace = get_namespace(root)
    distName_init = root.findall('*')[1].get('distName').split('/')[0]

    #removing all MOs
    et.strip_elements(tree,namespace+'managedObject')

    
    
    for sheet_name in sheet_names:
        sheet = pd.read_excel(updated_filter,sheet_name=sheet_name)

        for row in range(len(sheet)):
            
            # getting operation and its column location 
            operation = sheet.loc[row,'operation']
            class_ind = sheet.columns.get_loc("operation")
            dist_list = []
            
            # creating distName acording to given class' hierarchy
            for i in range(class_ind):
                dist_list.append('-'.join([sheet.columns[i],str(sheet.iloc[row,i])]))

            distName='/'.join([distName_init]+dist_list)

            # creating an MO 
            mo = et.SubElement(root,namespace+'managedObject',attrib={'operation':operation,'distName':distName,'class':sheet.columns[class_ind-1]})
            
            # iterating through each parameter
            for param,value in sheet.iloc[row,class_ind+1:].items():
                if type(value) != str:
                    if math.isnan(value):
                        continue
                    
                # if parameter is part of a list
                if len(param.split(':'))==3:

                    listName = param.split(':')[0]
                    item = param.split(':')[1].split('item')[1]
                    par = param.split(':')[2]
                    
                    
                    l = mo.find(namespace+'list[@name="%s"]'%listName)
                    # if list not already in the MO
                    if l==None:
                        l = et.SubElement(mo,namespace+'list',attrib={'name':listName})

                    if l.find(namespace+'item[@num="%s"]'%item)!=None:
                              et.SubElement(l.find(namespace+'item[@num="%s"]'%item),namespace+'p',attrib={'name': par }).text=str(value)
                    else:
                              item = et.SubElement(l,namespace+'item',attrib={'num':item})
                              et.SubElement(item,namespace+'p',attrib={'name': par }).text=str(value)
                else:
                    et.SubElement(mo,namespace+'p',attrib={'name': param }).text=str(value)

            for item in mo.findall('.//'+namespace+'item'):
                              item.attrib.pop("num",None)
    
    
    root = etree.ElementTree(tree.getroot())
    root.write('app/download/download.xml', pretty_print=True)

def clear_uploads(path):
    for file in os.listdir(path):
        os.remove(path+file)
    
@app.route('/')
@app.route('/index')
def index():
    return render_template("public/index.html")


@app.route("/upload-file.html", methods=["GET", "POST"])
def upload_file():
    upload()
    return render_template("public/upload-file.html")

            
@app.route('/result',methods = ['POST'])
def result():
    if request.method == 'POST':
        doc = xmlDocument+'sample.XML'
        print(doc)
        df = xml_to_dataframe(doc)
        params = request.form.to_dict()
        # filtering for class
        data = df[df['class']==params['class']]
        cl = params['class'] # 1. class
        id_ = params['class_id'] # 2.  id for class
        #filtering for id
        ids =  data.distName.str.split(cl+'-').str[1].str.split('/').str[0].str.strip()
        data = data[ids==id_]
        param = params['param']
        print(params)
        if param=='':
            param = 'All'
        else:
            data = data[data.parameter==param]
        header = list(data.columns)
        values = data.values
        table = data.to_html()
        text_file = open("app/templates/public/result.html", "w") 
        text_file.write(table)
        text_file.close() 
        data.to_csv("app/download/download.csv", index=False)
        return render_template("public/result.html",cl = cl, id_=id_,param=param,header = header, values=values, table=table)

@app.route('/xml-view',methods = ['POST'])
def xmlview():
    doc = xmlDocument+'sample.XML'
    params = request.form.to_dict()
    class_=params.get('class_').strip().lower()
    site_id=params.get('site_id').split(',')
    param_=params.get('param_').split(',')
    values=params.get('values').split(',')
    param_dict= {param_[i].strip().lower(): values[i] for i in range(len(param_))}
    updateXML(doc,class_,site_id,param_dict)
    return render_template("public/xml-view.html", class_=class_, site_id=site_id, param_=param_, values=values, param_dict=param_dict)

@app.route('/update.html',methods=["GET", "POST"])
def update():
    upload()
    return render_template("public/update.html")

@app.route('/heavy_update.html',methods=["GET", "POST"])
def heavy_update():
    upload()
    return render_template("public/heavy_update.html")

@app.route('/bulk_process.html', methods = ['POST'])
def bulk_process():
    doc = xmlDocument+'sample.XML'
    input_format = xmlDocument+'input.xlsx'
    input_xlsx= xmlDocument+'dump.xlsx'
    dumpparser(doc)
    filter_dump(input_format, input_xlsx)
    return render_template('public/heavy_update.html')

@app.route('/process_xml.html',methods=["GET", "POST"])
def process_xml():
    upload()
    flash("XML file uploaded!")
    return render_template("public/process_xml.html")

@app.route('/final_xml.html', methods = ["GET", 'POST'])
def final_xml():
    #Returning final XML
    doc = xmlDocument+'sample.XML'
    updated_filter = xmlDocument+'update_format.xlsx'
    create_XML(updated_filter, doc)
    return render_template('public/final_xml.html')

@app.route('/download/download.csv', methods=["GET"])
def plot_csv():
    path = 'app/download/download.csv'
    return_data = io.BytesIO()
    with open(path, 'rb') as fo:
        return_data.write(fo.read())
    # (after writing, cursor will be at last byte, so move it to start)
    return_data.seek(0)
    os.remove(path)
    clear_uploads('instance/uploads/')
    print("File Cleared!")
    return send_file(return_data,
                     mimetype='text/csv',
                     attachment_filename='result.csv',
                     as_attachment=True)

@app.route('/download/download.xml', methods=["GET"])
def plot_xml():
    path = 'app/download/download.xml'
    return_data = io.BytesIO()
    with open(path, 'rb') as fo:
        return_data.write(fo.read())
    # (after writing, cursor will be at last byte, so move it to start)
    return_data.seek(0)
    os.remove(path)
    clear_uploads('instance/uploads/')
    print("File Cleared!")
    return send_file(return_data,
                     mimetype='text/xml',
                     attachment_filename='result.xml',
                     as_attachment=True)

@app.route('/download/download.xlsx', methods=["GET"])
def plot_xlsx():
    path = 'app/download/download.xlsx'
    return_data = io.BytesIO()
    with open(path, 'rb') as fo:
        return_data.write(fo.read())
    # (after writing, cursor will be at last byte, so move it to start)
    return_data.seek(0)
    os.remove(path)
    clear_uploads('instance/uploads/')
    print("File Cleared!")
    return send_file(return_data,
                     mimetype='text/xlsx',
                     attachment_filename='result.xlsx',
                     as_attachment=True)

'''
@app.route('/download/update.xlsx', methods=["GET"])
def update_xlsx():
    return send_file('download/update.xlsx',
                     mimetype='text/xlsx',
                     attachment_filename='update.xlsx',
                     as_attachment=True)
@app.route('/bulk_process.html', methods = ['POST'])
def bulk_process():
    doc = xmlDocument+'sample.XML'
    inputDocument=xmlDocument+'Input.csv'
    bulkupdateXML(doc,inputDocument)
    return render_template('public/heavy_update.html')

    d = open(download_option, "r")
    download = str(d.read())
    text_file = open("app/templates/public/final_xml.html", "w") 
    text_file.write(download)
    text_file.close() 
'''
